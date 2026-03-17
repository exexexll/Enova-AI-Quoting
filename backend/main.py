"""FastAPI application entry point for the Enova AI Quoting System."""
from __future__ import annotations

import json
import logging
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

import aiofiles
from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from sse_starlette.sse import EventSourceResponse

from backend.config import (
    CLIENT_UPLOADS_DIR, EXPORTS_DIR, CLIENT_RECORDS_DIR,
    CONTRACTS_DIR, INGREDIENT_IMAGES_DIR, DATA_DIR, CORS_ORIGINS,
)
from backend.models.database import init_db, get_db, get_config, set_config
from backend.models.schemas import (
    SessionCreate, SessionOut, ChatRequest, ChatMessageOut,
    IngredientSearchResult, EscalationOut, EscalationCreate, EscalationResolve,
    ContractOut, ContractSubmit, ConfigItem, ImportResult, FileUploadOut,
)
from backend.services.excel_import import import_all
from backend.services.session_service import get_session, list_sessions, get_chat_history
from backend.services.excel_export import export_sample_request, export_client_record
from backend.services.contract_service import generate_contract
from backend.retrieval.indexer import build_all_indices
from backend.retrieval.hybrid_search import hybrid_search
from backend.agents.orchestrator import create_session, stream_response

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(name)s] %(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

from contextlib import asynccontextmanager

def _deduplicate_ingredients():
    """Remove duplicate ingredients (same item_name + source_tab), keeping the one with the lowest id."""
    with get_db() as conn:
        conn.execute("""
            DELETE FROM ingredients WHERE id NOT IN (
                SELECT MIN(id) FROM ingredients GROUP BY item_name, source_tab, supplier
            )
        """)
        remaining = conn.execute("SELECT COUNT(*) as cnt FROM ingredients").fetchone()
        return remaining["cnt"] if remaining else 0


@asynccontextmanager
async def lifespan(application: FastAPI):
    """Startup/shutdown lifecycle."""
    logger.info("Initializing database...")
    init_db()

    with get_db() as conn:
        count = conn.execute("SELECT COUNT(*) as cnt FROM ingredients").fetchone()
        already_imported = (count["cnt"] if count else 0) > 0

    if already_imported:
        logger.info("Ingredients already in DB (%d rows), skipping import.", count["cnt"] if count else 0)
        remaining = _deduplicate_ingredients()
        logger.info("After dedup: %d ingredients", remaining)
    else:
        try:
            logger.info("Importing ingredient data...")
            results = import_all()
            logger.info("Import results: %s", results)
        except Exception as e:
            logger.error("Ingredient import failed (non-fatal): %s", e)

    try:
        logger.info("Building search indices (BM25 only, embeddings on demand)...")
        build_all_indices(skip_embeddings=True)
    except Exception as e:
        logger.error("Index building failed (non-fatal): %s", e)
    logger.info("Startup complete!")
    yield
    logger.info("Shutting down...")

app = FastAPI(title="Enova AI Quoting System", version="1.0.0", lifespan=lifespan)

_KNOWN_ORIGINS = [
    "https://enova-ai-quoting-dcsp.vercel.app",
    "https://enova-ai-quoting-3wba.vercel.app",
    "https://officallol.vercel.app",
    "http://localhost:3000",
    "http://localhost:3001",
]
_env_origins = [o.strip() for o in CORS_ORIGINS.split(",") if o.strip()] if CORS_ORIGINS else []
_cors_origins = list(set(_env_origins + _KNOWN_ORIGINS)) if CORS_ORIGINS != "*" else ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials="*" not in _cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
async def root():
    """Root: point to API docs and health."""
    return {
        "app": "Enova AI Quoting System",
        "version": "1.0.0",
        "docs": "/docs",
        "health": "/api/health",
    }


@app.get("/api/health")
async def health_check():
    """Lightweight health check for monitoring and load balancers."""
    return {"status": "ok", "version": "1.0.0"}

# Serve static ingredient images
if INGREDIENT_IMAGES_DIR.exists():
    app.mount("/static/images", StaticFiles(directory=str(INGREDIENT_IMAGES_DIR)), name="images")


# ==================== SESSIONS ====================

@app.post("/api/sessions", response_model=SessionOut)
async def api_create_session(req: SessionCreate):
    session_id = await create_session(
        client_name=req.client_name,
        pre_selected_ingredient=req.pre_selected_ingredient,
    )
    return get_session(session_id)


@app.get("/api/sessions", response_model=list[SessionOut])
async def api_list_sessions(status: Optional[str] = None, limit: int = 50):
    return list_sessions(status=status, limit=min(limit, 500))


@app.get("/api/sessions/{session_id}", response_model=SessionOut)
async def api_get_session(session_id: str):
    s = get_session(session_id)
    if not s:
        raise HTTPException(404, "Session not found")
    return s


@app.get("/api/sessions/{session_id}/messages", response_model=list[ChatMessageOut])
async def api_get_messages(session_id: str):
    return get_chat_history(session_id)


# ==================== CHAT (SSE STREAMING) ====================

@app.post("/api/chat")
async def api_chat(req: ChatRequest):
    """Stream a chat response with thinking/executing tags via SSE."""
    # Validate session exists
    if not get_session(req.session_id):
        raise HTTPException(404, "Session not found")

    async def event_generator():
        async for event in stream_response(req.session_id, req.message):
            yield {
                "event": event["event"],
                "data": json.dumps({"content": event["data"]}) if isinstance(event["data"], str) else event["data"],
            }

    return EventSourceResponse(event_generator())


# ==================== INGREDIENT SEARCH ====================

@app.get("/api/ingredients/sources")
async def api_ingredient_sources():
    """Return distinct source_tab values for filter dropdowns."""
    with get_db() as conn:
        rows = conn.execute("SELECT DISTINCT source_tab FROM ingredients WHERE source_tab IS NOT NULL ORDER BY source_tab").fetchall()
    return [r["source_tab"] for r in rows]


@app.get("/api/ingredients/search", response_model=list[IngredientSearchResult])
async def api_search_ingredients(q: str, top_k: int = 10, source: Optional[str] = None):
    return hybrid_search(q, top_k=top_k, source_filter=source, use_embeddings=False)


@app.get("/api/ingredients")
async def api_list_ingredients(
    page: int = 1,
    per_page: int = 50,
    source: Optional[str] = None,
    has_price: Optional[bool] = None,
):
    """List ingredients with pagination."""
    offset = (page - 1) * per_page
    conditions = []
    params: list = []

    if source:
        conditions.append("source_tab = ?")
        params.append(source)
    if has_price is True:
        conditions.append("(sum_cavg > 0 OR cost_kg > 0 OR price_per_kg > 0)")
    elif has_price is False:
        conditions.append("needs_manual_price = 1")

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    with get_db() as conn:
        total = conn.execute(f"SELECT COUNT(*) as cnt FROM ingredients {where}", params).fetchone()["cnt"]
        rows = conn.execute(
            f"SELECT * FROM ingredients {where} ORDER BY item_name LIMIT ? OFFSET ?",
            params + [per_page, offset],
        ).fetchall()

    # Exclude binary embedding blob from API response
    items_out = []
    for r in rows:
        d = dict(r)
        d.pop("embedding", None)
        items_out.append(d)

    return {
        "items": items_out,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }


# ==================== FILE UPLOAD ====================

@app.post("/api/sessions/{session_id}/upload")
async def api_upload_file(session_id: str, file: UploadFile = File(...)):
    """Upload a file (image, PDF, Excel) for AI extraction."""
    # Validate session exists
    if not get_session(session_id):
        raise HTTPException(404, "Session not found")

    session_dir = CLIENT_UPLOADS_DIR / session_id / "original"
    session_dir.mkdir(parents=True, exist_ok=True)

    # Sanitize filename to prevent path traversal
    safe_name = re.sub(r'[^\w\-. ]', '_', Path(file.filename or "upload").name)
    if not safe_name or safe_name.startswith('.'):
        safe_name = f"upload_{int(datetime.now().timestamp())}"

    filepath = session_dir / safe_name
    async with aiofiles.open(str(filepath), "wb") as f:
        content = await file.read()
        await f.write(content)

    from backend.services.file_extraction import extract_file_content
    extracted = extract_file_content(str(filepath), file.content_type)

    with get_db() as conn:
        file_id = conn.execute_returning_id(
            """INSERT INTO client_files (session_id, filename, content_type, file_path, extraction_json)
               VALUES (?,?,?,?,?)""",
            (session_id, file.filename, file.content_type, str(filepath), extracted or None),
        )

    return {
        "id": file_id,
        "filename": file.filename,
        "path": str(filepath),
        "extracted": extracted[:500] if extracted else None,
        "extracted_length": len(extracted) if extracted else 0,
    }


@app.get("/api/sessions/{session_id}/files")
async def api_list_session_files(session_id: str):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM client_files WHERE session_id=? ORDER BY created_at",
            (session_id,),
        ).fetchall()
    return [dict(r) for r in rows]


# ==================== ESCALATION QUEUE ====================

@app.get("/api/escalations", response_model=list[EscalationOut])
async def api_list_escalations(status: str = "pending"):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM escalation_queue WHERE status=? ORDER BY created_at DESC",
            (status,),
        ).fetchall()
    return [
        EscalationOut(
            id=r["id"], session_id=r["session_id"], client_name=r["client_name"],
            item_requested=r["item_requested"], source=r["source"],
            quantity_needed=r["quantity_needed"], similar_items=r["similar_items"],
            est_low=r["est_low"], est_high=r["est_high"], status=r["status"],
            confirmed_price=r["confirmed_price"], admin_notes=r["admin_notes"],
            created_at=r["created_at"],
        )
        for r in rows
    ]


@app.post("/api/escalations", response_model=EscalationOut)
async def api_create_escalation(req: EscalationCreate):
    """Create an escalation from the client portal (e.g. missing price inquiry)."""
    client_name = None
    if req.session_id:
        session = get_session(req.session_id)
        if session:
            client_name = session.get("client_name") if isinstance(session, dict) else getattr(session, "client_name", None)
    with get_db() as conn:
        cursor = conn.execute(
            """INSERT INTO escalation_queue (session_id, client_name, item_requested, source, quantity_needed, similar_items)
               VALUES (?,?,?,?,?,?)""",
            (req.session_id, client_name, req.item_requested, req.source, req.quantity_needed, req.similar_items),
        )
        row = conn.execute("SELECT * FROM escalation_queue WHERE id=?", (cursor.lastrowid,)).fetchone()
    return EscalationOut(
        id=row["id"], session_id=row["session_id"], client_name=row["client_name"],
        item_requested=row["item_requested"], source=row["source"],
        quantity_needed=row["quantity_needed"], similar_items=row["similar_items"],
        est_low=row["est_low"], est_high=row["est_high"], status=row["status"],
        confirmed_price=row["confirmed_price"], admin_notes=row["admin_notes"],
        created_at=row["created_at"],
    )


@app.put("/api/escalations/{escalation_id}/resolve")
async def api_resolve_escalation(escalation_id: int, req: EscalationResolve):
    with get_db() as conn:
        conn.execute(
            """UPDATE escalation_queue
               SET status='resolved', confirmed_price=?, admin_notes=?, resolved_at=datetime('now')
               WHERE id=?""",
            (req.confirmed_price, req.admin_notes, escalation_id),
        )
    return {"status": "resolved", "id": escalation_id}


# ==================== CONTRACTS ====================

@app.post("/api/sessions/{session_id}/contract")
async def api_generate_contract(session_id: str):
    path = generate_contract(session_id)
    return {"path": path, "session_id": session_id}


@app.get("/api/sessions/{session_id}/contract/download")
async def api_download_contract(session_id: str):
    with get_db() as conn:
        row = conn.execute(
            "SELECT pdf_path FROM contracts WHERE session_id=? ORDER BY version DESC LIMIT 1",
            (session_id,),
        ).fetchone()
    if not row or not row["pdf_path"]:
        raise HTTPException(404, "No contract found")
    pdf_path = Path(row["pdf_path"]).resolve()
    if not str(pdf_path).startswith(str(CONTRACTS_DIR.resolve())):
        raise HTTPException(403, "Access denied")
    if not pdf_path.exists():
        raise HTTPException(404, "Contract file not found on disk")
    return FileResponse(str(pdf_path), filename=pdf_path.name)


@app.post("/api/sessions/{session_id}/contract/submit")
async def api_submit_contract(session_id: str, req: ContractSubmit):
    with get_db() as conn:
        conn.execute(
            """UPDATE contracts SET status='submitted', client_name_sig=?, client_comments=?,
               submitted_at=datetime('now'), updated_at=datetime('now')
               WHERE session_id=? AND status IN ('draft', 'under_review')""",
            (req.client_name_sig, req.client_comments, session_id),
        )
        conn.execute(
            "UPDATE sessions SET contract_status='submitted', updated_at=datetime('now') WHERE id=?",
            (session_id,),
        )
    # Also export client record
    export_client_record(session_id)
    return {"status": "submitted", "session_id": session_id}


@app.get("/api/contracts")
async def api_list_contracts(status: Optional[str] = None):
    with get_db() as conn:
        if status:
            rows = conn.execute(
                "SELECT * FROM contracts WHERE status=? ORDER BY created_at DESC", (status,),
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM contracts ORDER BY created_at DESC").fetchall()
    return [dict(r) for r in rows]


@app.put("/api/contracts/{contract_id}/accept")
async def api_accept_contract(contract_id: int):
    with get_db() as conn:
        row = conn.execute("SELECT session_id FROM contracts WHERE id=?", (contract_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Contract not found")
        conn.execute(
            "UPDATE contracts SET status='accepted', accepted_at=datetime('now') WHERE id=?",
            (contract_id,),
        )
        conn.execute(
            "UPDATE sessions SET contract_status='accepted', status='completed' WHERE id=?",
            (row["session_id"],),
        )
    return {"status": "accepted"}


@app.put("/api/contracts/{contract_id}/revision")
async def api_request_revision(contract_id: int, notes: str = ""):
    with get_db() as conn:
        conn.execute(
            "UPDATE contracts SET status='revision', admin_notes=? WHERE id=?",
            (notes, contract_id),
        )
    return {"status": "revision"}


# ==================== SAMPLE ORDERS ====================

@app.get("/api/admin/sample-orders")
async def api_list_sample_orders():
    """List sessions in sample-related workflow states with their ingredients and quotes."""
    sample_states = ('sample_decision', 'sample_payment', 'sample_production', 'sample_confirmation')
    with get_db() as conn:
        sessions = conn.execute(
            f"SELECT * FROM sessions WHERE workflow_state IN ({','.join('?' * len(sample_states))}) "
            "ORDER BY updated_at DESC",
            sample_states,
        ).fetchall()

        orders = []
        for s in sessions:
            ingredients = conn.execute(
                "SELECT ingredient_name, mg_per_serving, confidence, cost_source "
                "FROM session_ingredients WHERE session_id=?",
                (s["id"],),
            ).fetchall()
            quote = conn.execute(
                "SELECT * FROM quotes WHERE session_id=? ORDER BY version DESC LIMIT 1",
                (s["id"],),
            ).fetchone()
            orders.append({
                "session_id": s["id"],
                "client_name": s["client_name"],
                "client_email": s["client_email"],
                "client_company": s["client_company"],
                "workflow_state": s["workflow_state"],
                "updated_at": s["updated_at"],
                "ingredients": [dict(i) for i in ingredients],
                "quote": dict(quote) if quote else None,
            })
    return orders


# ==================== EXPORT ====================

@app.post("/api/sessions/{session_id}/export/sample")
async def api_export_sample(session_id: str, notes: str = ""):
    path = export_sample_request(session_id, notes)
    return FileResponse(path, filename=Path(path).name)


@app.post("/api/sessions/{session_id}/export/record")
async def api_export_record(session_id: str):
    path = export_client_record(session_id)
    return FileResponse(path, filename=Path(path).name)


# ==================== ADMIN: CONFIG ====================

@app.get("/api/config")
async def api_get_config():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM pricing_config ORDER BY key").fetchall()
    return [dict(r) for r in rows]


@app.put("/api/config/{key}")
async def api_set_config(key: str, item: ConfigItem):
    set_config(key, item.value, item.description or "")
    return {"status": "updated", "key": key}


# ==================== ADMIN: DB IMPORTS ====================

@app.post("/api/admin/import/machine-rates", response_model=ImportResult)
async def api_import_machine_rates(file: UploadFile = File(...)):
    return await _import_rates_excel(file, "machine_rates", [
        "machine_type", "model", "hourly_rate", "setup_cost", "cleaning_cost",
        "throughput_per_hour", "maintenance_pct", "notes",
    ])


@app.post("/api/admin/import/labor-rates", response_model=ImportResult)
async def api_import_labor_rates(file: UploadFile = File(...)):
    return await _import_rates_excel(file, "labor_rates", [
        "role", "hourly_rate", "headcount_per_line", "est_hours_per_10k_units",
        "overtime_multiplier", "notes",
    ])


@app.post("/api/admin/import/packaging-rates", response_model=ImportResult)
async def api_import_packaging_rates(file: UploadFile = File(...)):
    return await _import_rates_excel(file, "packaging_rates", [
        "component_type", "description", "cost_per_unit", "min_order_qty",
        "lead_time_days", "supplier", "notes",
    ])


@app.post("/api/admin/import/transport-rates", response_model=ImportResult)
async def api_import_transport_rates(file: UploadFile = File(...)):
    return await _import_rates_excel(file, "transport_rates", [
        "carrier", "service_level", "rate_type", "rate_value",
        "weight_min_lbs", "weight_max_lbs", "zone_or_region", "surcharges_pct", "notes",
    ])


_ALLOWED_TABLES = {"machine_rates", "labor_rates", "packaging_rates", "transport_rates"}

async def _import_rates_excel(file: UploadFile, table_name: str, expected_cols: list[str]) -> ImportResult:
    """Generic Excel import for admin rate tables."""
    import openpyxl

    # Whitelist table names to prevent SQL injection
    if table_name not in _ALLOWED_TABLES:
        return ImportResult(rows_imported=0, errors=[f"Invalid table: {table_name}"])

    # Sanitize filename
    safe_fn = re.sub(r'[^\w\-. ]', '_', Path(file.filename or "import").name)
    save_path = DATA_DIR / "admin_imports" / safe_fn
    save_path.parent.mkdir(parents=True, exist_ok=True)
    content = await file.read()
    with open(str(save_path), "wb") as f:
        f.write(content)

    wb = openpyxl.load_workbook(str(save_path), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return ImportResult(rows_imported=0, errors=["Workbook has no active sheet"])

    # Read headers
    headers: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(c).strip().lower().replace(" ", "_") if c else "" for c in row]
        break

    errors: list[str] = []
    rows_imported = 0

    # Map columns — exact match first, then substring match
    col_map: dict[str, int] = {}
    for exp in expected_cols:
        exp_lower = exp.lower()
        # Try exact match first
        for i, h in enumerate(headers):
            if h == exp_lower:
                col_map[exp] = i
                break
        else:
            # Fallback: substring match (exp contained in header or header contained in exp)
            for i, h in enumerate(headers):
                if h and (exp_lower in h or h in exp_lower) and i not in col_map.values():
                    col_map[exp] = i
                    break

    with get_db() as conn:
        conn.execute(f"DELETE FROM {table_name}")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            values = {}
            for col_name, col_idx in col_map.items():
                if col_idx < len(row):
                    values[col_name] = row[col_idx]

            if not values:
                continue

            cols = ", ".join(values.keys())
            placeholders = ", ".join("?" * len(values))
            try:
                conn.execute(
                    f"INSERT INTO {table_name} ({cols}) VALUES ({placeholders})",
                    list(values.values()),
                )
                rows_imported += 1
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

    wb.close()
    return ImportResult(rows_imported=rows_imported, errors=errors)


# ==================== ADMIN: RATES LISTING ====================

@app.get("/api/admin/machine-rates")
async def api_list_machine_rates():
    with get_db() as conn:
        return [dict(r) for r in conn.execute("SELECT * FROM machine_rates").fetchall()]


@app.get("/api/admin/labor-rates")
async def api_list_labor_rates():
    with get_db() as conn:
        return [dict(r) for r in conn.execute("SELECT * FROM labor_rates").fetchall()]


@app.get("/api/admin/packaging-rates")
async def api_list_packaging_rates():
    with get_db() as conn:
        return [dict(r) for r in conn.execute("SELECT * FROM packaging_rates").fetchall()]


@app.get("/api/admin/transport-rates")
async def api_list_transport_rates():
    with get_db() as conn:
        return [dict(r) for r in conn.execute("SELECT * FROM transport_rates").fetchall()]


# ==================== INGREDIENT IMAGES (SerpAPI) ====================

_image_cache: dict[str, str] = {}

@app.get("/api/ingredient-image")
async def api_ingredient_image(name: str):
    """Get an image URL for an ingredient using SerpAPI Google Images."""
    from backend.config import SERPAPI_KEY

    if name in _image_cache:
        return {"url": _image_cache[name]}

    try:
        from serpapi import GoogleSearch
        params = {
            "engine": "google_images",
            "q": f"{name} supplement powder ingredient",
            "api_key": SERPAPI_KEY,
            "num": 1,
        }
        search = GoogleSearch(params)
        results = search.get_dict()
        images = results.get("images_results", [])
        if images:
            url = images[0].get("thumbnail", images[0].get("original", ""))
            _image_cache[name] = url
            return {"url": url}
    except Exception as e:
        logger.warning("SerpAPI image search failed for '%s': %s", name, e)

    return {"url": ""}


# ==================== REFRESH INDICES ====================

@app.post("/api/admin/refresh-indices")
async def api_refresh_indices(include_embeddings: bool = False):
    import_all()
    results = build_all_indices(skip_embeddings=not include_embeddings)
    return {"status": "refreshed", **results}


@app.post("/api/admin/import/ingredient-master")
async def api_import_ingredient_master(file: UploadFile = File(...)):
    """Upload the Ingredient Master Excel and import all tabs."""
    safe_fn = re.sub(r'[^\w\-. ]', '_', Path(file.filename or "master.xlsx").name)
    save_path = DATA_DIR / "admin_imports" / safe_fn
    save_path.parent.mkdir(parents=True, exist_ok=True)
    content = await file.read()
    with open(str(save_path), "wb") as f:
        f.write(content)

    from backend.services.excel_import import import_enova_data, import_master_tab, import_supplier_tab
    results: dict[str, int] = {}
    results["enova_data"] = import_enova_data(save_path)
    results["master"] = import_master_tab(save_path)
    supplier_tabs = [
        ("UAA US", "uaa_us"), ("Kingdom US", "kingdom_us"),
        ("CharlesBowman", "charlesbowman"), ("LGB", "lgb"),
        ("Nutravative", "nutravative"), ("专利原料", "patented"),
        ("Runde Glucosamine", "runde"),
    ]
    for tab_name, source_key in supplier_tabs:
        results[source_key] = import_supplier_tab(save_path, tab_name, source_key)

    build_all_indices(skip_embeddings=True)
    total = sum(results.values())
    return {"status": "imported", "total": total, "tabs": results}
