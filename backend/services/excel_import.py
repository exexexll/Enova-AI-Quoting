"""Import ingredient data from the Ingredient Master Excel file into SQLite."""
from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from backend.config import INGREDIENT_MASTER_PATH
from backend.models.database import get_db

logger = logging.getLogger(__name__)


def _safe_float(val, default=0.0) -> float:
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _safe_str(val, default="") -> str:
    if val is None:
        return default
    s = str(val).strip()
    # Clean up tab/newline chars from messy Excel data
    s = s.replace("\t", " ").replace("\n", " ").replace("\r", " ")
    while "  " in s:
        s = s.replace("  ", " ")
    return s


def import_enova_data(filepath: Path | None = None) -> int:
    """Import the 'Enova Data' tab into the ingredients table.

    Returns number of rows imported.
    """
    filepath = filepath or INGREDIENT_MASTER_PATH
    if not filepath.exists():
        logger.warning("Ingredient Master file not found at %s", filepath)
        return 0

    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    ws = wb["Enova Data"]

    rows_imported = 0
    with get_db() as conn:
        # Clear existing enova_data entries (temporarily disable FK to avoid cascade issues
        # with session_ingredients referencing these IDs from prior sessions)
        conn.execute("PRAGMA foreign_keys=OFF")
        conn.execute("DELETE FROM ingredients WHERE source_tab='enova_data'")

        batch: list[tuple] = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if row[0] is None:
                continue

            item_name = _safe_str(row[0])
            if not item_name:
                continue

            supplier = _safe_str(row[1])
            if supplier == "(blank)":
                supplier = ""
            location = _safe_str(row[2])
            uom = _safe_str(row[3])
            c_last = _safe_float(row[4])
            ss = _safe_float(row[5])
            on_hand = _safe_float(row[6])
            sum_cavg = _safe_float(row[7])
            sum_ss_cost = _safe_float(row[8])
            cost_kg = _safe_float(row[9])
            supplier_code = _safe_str(row[10]) if len(row) > 10 else ""

            needs_manual = 1 if (sum_cavg == 0 and sum_ss_cost == 0 and cost_kg == 0) else 0

            batch.append((
                item_name, None, supplier, location, uom,
                c_last, ss, on_hand, sum_cavg, sum_ss_cost, cost_kg,
                supplier_code, "enova_data", None,
                None, None, None, 0, None, None, None,
                needs_manual,
            ))

            if len(batch) >= 500:
                conn.executemany(
                    """INSERT INTO ingredients
                       (item_name, item_id, supplier, location, uom,
                        c_last, safety_stock, on_hand, sum_cavg, sum_ss_cost, cost_kg,
                        supplier_code, source_tab, category,
                        chinese_name, potency, form, is_trademarked, warehouse, moq_kg, price_per_kg,
                        needs_manual_price)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    batch,
                )
                rows_imported += len(batch)
                batch.clear()

        if batch:
            conn.executemany(
                """INSERT INTO ingredients
                   (item_name, item_id, supplier, location, uom,
                    c_last, safety_stock, on_hand, sum_cavg, sum_ss_cost, cost_kg,
                    supplier_code, source_tab, category,
                    chinese_name, potency, form, is_trademarked, warehouse, moq_kg, price_per_kg,
                    needs_manual_price)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                batch,
            )
            rows_imported += len(batch)

    wb.close()
    logger.info("Imported %d items from Enova Data tab", rows_imported)
    return rows_imported


def import_master_tab(filepath: Path | None = None) -> int:
    """Import the 'Master' tab (supplier pricing reference)."""
    filepath = filepath or INGREDIENT_MASTER_PATH
    if not filepath.exists():
        return 0

    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    ws = wb["Master"]

    rows_imported = 0
    with get_db() as conn:
        conn.execute("PRAGMA foreign_keys=OFF")
        conn.execute("DELETE FROM ingredients WHERE source_tab='master'")

        batch: list[tuple] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue

            ingredient = _safe_str(row[0])
            chinese_name = _safe_str(row[1])
            potency = _safe_str(row[2])
            form = _safe_str(row[3])
            is_tm = 1 if _safe_str(row[4]).lower().startswith("yes") else 0
            supplier = _safe_str(row[5])
            warehouse = _safe_str(row[6])
            moq_kg = _safe_float(row[7])
            price_per_kg = _safe_float(row[8])

            batch.append((
                ingredient, None, supplier, None, "gm",
                0, 0, 0, 0, 0, 0,
                "", "master", None,
                chinese_name, potency, form, is_tm, warehouse, moq_kg, price_per_kg,
                0 if price_per_kg > 0 else 1,
            ))

        if batch:
            conn.executemany(
                """INSERT INTO ingredients
                   (item_name, item_id, supplier, location, uom,
                    c_last, safety_stock, on_hand, sum_cavg, sum_ss_cost, cost_kg,
                    supplier_code, source_tab, category,
                    chinese_name, potency, form, is_trademarked, warehouse, moq_kg, price_per_kg,
                    needs_manual_price)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                batch,
            )
            rows_imported = len(batch)

    wb.close()
    logger.info("Imported %d items from Master tab", rows_imported)
    return rows_imported


def import_supplier_tab(filepath: Path | None, tab_name: str, source_key: str) -> int:
    """Generic importer for additional supplier tabs (UAA US, Kingdom US, etc.)."""
    filepath = filepath or INGREDIENT_MASTER_PATH
    if not filepath.exists():
        return 0

    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    if tab_name not in wb.sheetnames:
        wb.close()
        return 0

    ws = wb[tab_name]
    rows_imported = 0

    with get_db() as conn:
        conn.execute("PRAGMA foreign_keys=OFF")
        conn.execute("DELETE FROM ingredients WHERE source_tab=?", (source_key,))

        batch: list[tuple] = []
        # Read header row to understand structure
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [_safe_str(c).lower() for c in row]
            break

        # Find likely name and price columns
        name_col = None
        price_col = None
        for i, h in enumerate(headers):
            if any(kw in h for kw in ["product name", "description", "ingredient", "品名"]):
                name_col = i
            if any(kw in h for kw in ["price", "单价", "cost"]):
                price_col = i

        # Fallback: first column is name
        if name_col is None:
            # Try column 0 or 1
            name_col = 1 if len(headers) > 1 and not headers[0] else 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[name_col] is None:
                continue

            item_name = _safe_str(row[name_col])
            if not item_name or len(item_name) < 2:
                continue

            price = _safe_float(row[price_col]) if price_col is not None and price_col < len(row) else 0

            batch.append((
                item_name, None, "", None, "gm",
                0, 0, 0, 0, 0, 0,
                "", source_key, None,
                None, None, None, 0, None, None, price,
                0 if price > 0 else 1,
            ))

        if batch:
            conn.executemany(
                """INSERT INTO ingredients
                   (item_name, item_id, supplier, location, uom,
                    c_last, safety_stock, on_hand, sum_cavg, sum_ss_cost, cost_kg,
                    supplier_code, source_tab, category,
                    chinese_name, potency, form, is_trademarked, warehouse, moq_kg, price_per_kg,
                    needs_manual_price)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                batch,
            )
            rows_imported = len(batch)

    wb.close()
    logger.info("Imported %d items from %s tab", rows_imported, tab_name)
    return rows_imported


def import_all() -> dict[str, int]:
    """Import all tabs from the Ingredient Master file."""
    results: dict[str, int] = {}
    results["enova_data"] = import_enova_data()
    results["master"] = import_master_tab()

    supplier_tabs = [
        ("UAA US", "uaa_us"),
        ("Kingdom US", "kingdom_us"),
        ("CharlesBowman", "charlesbowman"),
        ("LGB", "lgb"),
        ("Nutravative", "nutravative"),
        ("专利原料", "patented"),
        ("Runde Glucosamine", "runde"),
    ]
    for tab_name, source_key in supplier_tabs:
        results[source_key] = import_supplier_tab(None, tab_name, source_key)

    total = sum(results.values())
    logger.info("Total imported: %d items across %d tabs", total, len(results))
    return results
