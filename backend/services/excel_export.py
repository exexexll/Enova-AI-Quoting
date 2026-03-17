"""Excel export services for R&D sample requests and client info recording."""
from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from backend.config import EXPORTS_DIR, CLIENT_RECORDS_DIR
from backend.models.database import get_db

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")


def _add_header_row(ws, headers: list[str], row: int = 1):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT_WHITE
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


_SECTION_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_SECTION_FONT = Font(bold=True, size=12, color="FFFFFF")
_LABEL_FONT = Font(bold=True, size=10)
_NOTE_FONT = Font(italic=True, size=9, color="666666")


def _add_section(ws, row: int, title: str, cols: int = 2) -> int:
    """Add a blue section header spanning cols columns. Returns next row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    cell.alignment = Alignment(horizontal="left")
    return row + 1


def _add_field(ws, row: int, label: str, value, note: str = "") -> int:
    """Add a label-value pair. Returns next row."""
    ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
    ws.cell(row=row, column=2, value=value or "")
    if note:
        ws.cell(row=row, column=3, value=note).font = _NOTE_FONT
    return row + 1


def export_sample_request(session_id: str, special_notes: str = "") -> str:
    """Export a comprehensive R&D sample request Excel.

    Includes everything the lab needs: client info, product specs,
    full formulation, batch calculations, packaging, QC requirements,
    and manufacturing instructions.
    """
    with get_db() as conn:
        session = conn.execute("SELECT * FROM sessions WHERE id=?", (session_id,)).fetchone()
        ingredients = conn.execute(
            "SELECT * FROM session_ingredients WHERE session_id=?", (session_id,),
        ).fetchall()
        quote = conn.execute(
            "SELECT * FROM quotes WHERE session_id=? ORDER BY version DESC LIMIT 1",
            (session_id,),
        ).fetchone()
        messages = conn.execute(
            "SELECT * FROM chat_messages WHERE session_id=? AND role='user' ORDER BY timestamp",
            (session_id,),
        ).fetchall()

    ctx = json.loads(session["context_json"]) if session and session["context_json"] else {}
    specs = ctx.get("product_specs", {})

    product_type = specs.get("product_type", "capsule")
    serving_size = specs.get("serving_size", 2)
    servings_per_unit = specs.get("servings_per_unit", 90)
    total_count = specs.get("total_count", serving_size * servings_per_unit)
    order_qty = specs.get("order_quantity", 10000)

    total_active_mg = sum(
        (ing["mg_per_serving"] or 0) for ing in ingredients
    )

    wb = openpyxl.Workbook()

    # ==================== Sheet 1: Sample Request Summary ====================
    ws = wb.active or wb.create_sheet("Sample Request")
    ws.title = "Sample Request"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 35

    r = 1
    r = _add_section(ws, r, "ENOVA SCIENCE — R&D SAMPLE REQUEST", 3)
    r = _add_field(ws, r, "Request ID", session_id)
    r = _add_field(ws, r, "Request Date", datetime.now().strftime("%Y-%m-%d %H:%M"))
    r = _add_field(ws, r, "Priority", "Standard", "Rush = +$500, 5 business days")
    r = _add_field(ws, r, "Status", session["workflow_state"] if session else "")
    r += 1

    r = _add_section(ws, r, "CLIENT INFORMATION", 3)
    r = _add_field(ws, r, "Client Name", session["client_name"] if session else "")
    r = _add_field(ws, r, "Company", session["client_company"] if session else "")
    r = _add_field(ws, r, "Email", session["client_email"] if session else "")
    r = _add_field(ws, r, "Phone", session["client_phone"] if session else "")
    r = _add_field(ws, r, "Address", session["client_address"] if session else "")
    r += 1

    r = _add_section(ws, r, "PRODUCT SPECIFICATION", 3)
    r = _add_field(ws, r, "Product Name", specs.get("product_name", "TBD"))
    r = _add_field(ws, r, "Dosage Form", product_type.title(), "Capsule / Tablet / Powder / Gummy / Liquid / Softgel")
    r = _add_field(ws, r, "Capsule Type", specs.get("capsule_type", "HPMC Vegetarian 00"), "HPMC / Gelatin / Size 000-4")
    r = _add_field(ws, r, "Capsule Size", "00", "000=1000mg, 00=735mg, 0=500mg, 1=400mg")
    r = _add_field(ws, r, "Serving Size", f"{serving_size} {product_type}(s)")
    r = _add_field(ws, r, "Servings Per Container", servings_per_unit)
    r = _add_field(ws, r, "Total Count Per Unit", total_count)
    r = _add_field(ws, r, "Total Active Fill (mg/serving)", round(total_active_mg, 1))
    r = _add_field(ws, r, "Target Fill Weight (mg/capsule)", round(total_active_mg / serving_size, 1) if serving_size > 0 else "", "Active + excipients")
    r += 1

    r = _add_section(ws, r, "BATCH & ORDER DETAILS", 3)
    r = _add_field(ws, r, "Sample Type", "Pilot/Sample Run", "Bench (10-50 units) / Pilot (500-2000 units)")
    r = _add_field(ws, r, "Sample Quantity", "2,000 bottles", "Adjust as needed")
    r = _add_field(ws, r, "Production MOQ", f"{order_qty:,} units")
    r = _add_field(ws, r, "Batch Size (capsules)", f"{total_count * 2000:,}", "= total_count × sample_qty")
    r = _add_field(ws, r, "Expected Yield", "95-98%", "Account for 2-5% waste")
    r += 1

    r = _add_section(ws, r, "PACKAGING REQUIREMENTS", 3)
    r = _add_field(ws, r, "Primary Container", "120cc HDPE Bottle", "White round / Amber glass / Custom")
    r = _add_field(ws, r, "Closure", "CRC Cap (child-resistant)", "CRC / Flip-top / Screw cap")
    r = _add_field(ws, r, "Seal", "Induction seal", "Induction / Shrink band / None")
    r = _add_field(ws, r, "Desiccant", "Silica gel canister", "Required for hygroscopic formulas")
    r = _add_field(ws, r, "Label", "Full-color wrap label", "For sample: plain/simplified OK")
    r = _add_field(ws, r, "Cotton/Rayon", "Yes", "Filler for bottle headspace")
    r = _add_field(ws, r, "Outer Case", "24 bottles/case", "Corrugated shipper")
    r += 1

    r = _add_section(ws, r, "QUALITY & COMPLIANCE", 3)
    r = _add_field(ws, r, "GMP Compliance", "21 CFR Part 111", "Dietary supplement cGMP")
    r = _add_field(ws, r, "Certifications Needed", "N/A", "Organic / NSF / Non-GMO / Kosher / Halal")
    r = _add_field(ws, r, "Allergen Statement", "Check formulation", "Soy / Dairy / Gluten / Shellfish / Tree nuts")
    r = _add_field(ws, r, "Testing Required", "Identity, Potency, Micro, Heavy Metals", "Per USP <2091>")
    r = _add_field(ws, r, "Stability Testing", "Accelerated (40°C/75%RH, 3 months)", "If required")
    r = _add_field(ws, r, "Shelf Life Target", "24 months", "From date of manufacture")
    r = _add_field(ws, r, "Storage Conditions", "Room temperature, dry place", "Below 25°C / 77°F")
    r += 1

    r = _add_section(ws, r, "SPECIAL NOTES", 3)
    r = _add_field(ws, r, "Client Notes", special_notes or "None")
    r = _add_field(ws, r, "R&D Notes", "", "To be filled by R&D team")

    # ==================== Sheet 2: Formulation (Bill of Materials) ====================
    ws2 = wb.create_sheet("Formulation BOM")
    _add_header_row(ws2, [
        "Item #", "Ingredient Name", "Function", "mg/Serving",
        "mg/Capsule", "% of Fill", "Cost/Serving",
        "Confidence", "Source", "Supplier", "Notes",
    ])
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 10
    ws2.column_dimensions["G"].width = 12
    ws2.column_dimensions["H"].width = 12
    ws2.column_dimensions["I"].width = 12
    ws2.column_dimensions["J"].width = 18
    ws2.column_dimensions["K"].width = 20

    for idx, ing in enumerate(ingredients, 1):
        row = idx + 1
        mg = ing["mg_per_serving"] or 0
        mg_per_cap = mg / serving_size if serving_size > 0 else mg
        fill_pct = (mg / total_active_mg * 100) if total_active_mg > 0 else 0

        ws2.cell(row=row, column=1, value=idx)
        ws2.cell(row=row, column=2, value=ing["ingredient_name"])
        ws2.cell(row=row, column=3, value="Active")
        ws2.cell(row=row, column=4, value=round(mg, 2))
        ws2.cell(row=row, column=5, value=round(mg_per_cap, 2))
        ws2.cell(row=row, column=6, value=f"{fill_pct:.1f}%")
        ws2.cell(row=row, column=7, value=round(ing["unit_cost"] or 0, 6))
        ws2.cell(row=row, column=8, value=ing["confidence"] or "")
        ws2.cell(row=row, column=9, value=ing["cost_source"] or "")
        ws2.cell(row=row, column=10, value="")
        ws2.cell(row=row, column=11, value="")

    excipient_row = len(ingredients) + 2
    excipients = [
        ("Microcrystalline Cellulose (MCC)", "Filler / Bulking agent", "Q.S."),
        ("Silicon Dioxide", "Flow agent / Anti-caking", "1-2%"),
        ("Magnesium Stearate", "Lubricant", "0.5-1%"),
        ("Capsule Shell (HPMC)", "Capsule", f"{serving_size} ea/serving"),
    ]
    for i, (name, function, amount) in enumerate(excipients):
        row = excipient_row + i
        ws2.cell(row=row, column=1, value=len(ingredients) + i + 1)
        ws2.cell(row=row, column=2, value=name)
        ws2.cell(row=row, column=3, value=function)
        ws2.cell(row=row, column=4, value=amount)
        ws2.cell(row=row, column=11, value="Adjust per R&D")

    # Totals row
    total_row = excipient_row + len(excipients) + 1
    ws2.cell(row=total_row, column=2, value="TOTAL ACTIVE per serving").font = _LABEL_FONT
    ws2.cell(row=total_row, column=4, value=round(total_active_mg, 2)).font = _LABEL_FONT

    # ==================== Sheet 3: Manufacturing Instructions ====================
    ws3 = wb.create_sheet("Manufacturing")
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 50
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 25

    _add_header_row(ws3, ["Step", "Instruction", "Equipment", "QC Checkpoint"])
    steps = [
        ("1", "Receive and verify all raw materials against BOM. Check COAs.", "Receiving area", "Identity test each RM"),
        ("2", "Weigh all active ingredients per batch formula.", "Calibrated balance", "Record actual weights; ±2% tolerance"),
        ("3", "Weigh excipients (MCC, SiO₂, Mg stearate) per batch formula.", "Calibrated balance", "Record actual weights"),
        ("4", "Pre-blend active ingredients in V-blender for 10-15 min.", "V-Blender / Ribbon mixer", "Visual uniformity check"),
        ("5", "Add MCC filler, blend for additional 10 min.", "V-Blender", "Blend uniformity test (RSD <5%)"),
        ("6", "Add flow agents (SiO₂), blend 3-5 min. Add lubricant (Mg stearate), blend 2-3 min.", "V-Blender", "Flow test; angle of repose <40°"),
        ("7", "Encapsulate blend into capsule shells.", "Capsule filler (automatic)", "Fill weight check every 15 min; ±5% target weight"),
        ("8", "Polish filled capsules to remove dust.", "Capsule polisher", "Visual inspection: no cracks, dents, or leaks"),
        ("9", "Perform in-process QC: weight uniformity, disintegration.", "QC lab", "USP <2040> weight variation; <2091> disintegration <30 min"),
        ("10", "Count and fill into primary containers.", "Counting machine", "Verify count ± 0"),
        ("11", "Insert desiccant and cotton, apply induction seal.", "Sealing station", "Seal integrity check"),
        ("12", "Apply closure (CRC cap), label, and neckband.", "Labeling line", "Label accuracy and placement QC"),
        ("13", "Case pack and palletize. Record batch yield.", "Packaging area", "Final yield within 95-100% of theoretical"),
        ("14", "Submit retain samples and batch record to QA for release.", "QA", "Full batch documentation review"),
    ]
    for i, (step, instruction, equipment, qc) in enumerate(steps, 2):
        ws3.cell(row=i, column=1, value=step)
        ws3.cell(row=i, column=2, value=instruction)
        ws3.cell(row=i, column=3, value=equipment)
        ws3.cell(row=i, column=4, value=qc)

    # ==================== Sheet 4: Pricing Summary ====================
    ws4 = wb.create_sheet("Pricing")
    _add_header_row(ws4, ["Component", "Low", "Mid", "High", "Source"])
    if quote:
        rows = [
            ("Ingredients", quote["ingredient_cost_low"], quote["ingredient_cost_mid"], quote["ingredient_cost_high"], "DB"),
            ("Machine", quote["machine_cost_low"], quote["machine_cost_mid"], quote["machine_cost_high"], "Config"),
            ("Labor", quote["labor_cost_low"], quote["labor_cost_mid"], quote["labor_cost_high"], "Config"),
            ("Packaging", quote["packaging_cost_low"], quote["packaging_cost_mid"], quote["packaging_cost_high"], "Default"),
            ("Shipping", quote["transport_cost_low"], quote["transport_cost_mid"], quote["transport_cost_high"], "Rate table"),
            ("TOTAL (per unit)", quote["total_low"], quote["total_mid"], quote["total_high"],
             f"Margin: {(quote['margin_pct'] or 0) * 100:.0f}%"),
        ]
        for i, (comp, lo, mi, hi, src) in enumerate(rows, 2):
            ws4.cell(row=i, column=1, value=comp)
            ws4.cell(row=i, column=2, value=lo)
            ws4.cell(row=i, column=3, value=mi)
            ws4.cell(row=i, column=4, value=hi)
            ws4.cell(row=i, column=5, value=src)
        total_row = len(rows) + 3
        ws4.cell(row=total_row, column=1, value="Order Total").font = _LABEL_FONT
        ws4.cell(row=total_row, column=2, value=round((quote["total_low"] or 0) * order_qty, 2))
        ws4.cell(row=total_row, column=3, value=round((quote["total_mid"] or 0) * order_qty, 2))
        ws4.cell(row=total_row, column=4, value=round((quote["total_high"] or 0) * order_qty, 2))
        ws4.cell(row=total_row, column=5, value=f"{order_qty:,} units")

    # ==================== Sheet 5: Client Requirements ====================
    ws5 = wb.create_sheet("Client Requirements")
    _add_header_row(ws5, ["Timestamp", "Client Message"])
    ws5.column_dimensions["A"].width = 20
    ws5.column_dimensions["B"].width = 80
    for i, msg in enumerate(messages, 2):
        ws5.cell(row=i, column=1, value=msg["timestamp"])
        ws5.cell(row=i, column=2, value=msg["content"])

    filename = f"sample_request_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = EXPORTS_DIR / filename
    wb.save(str(filepath))
    logger.info("Exported sample request to %s", filepath)
    return str(filepath)


def export_client_record(session_id: str) -> str:
    """Export comprehensive client record Excel with all session data.

    Returns path to generated file.
    """
    with get_db() as conn:
        session = conn.execute("SELECT * FROM sessions WHERE id=?", (session_id,)).fetchone()
        ingredients = conn.execute(
            "SELECT * FROM session_ingredients WHERE session_id=?", (session_id,),
        ).fetchall()
        messages = conn.execute(
            "SELECT * FROM chat_messages WHERE session_id=? ORDER BY timestamp",
            (session_id,),
        ).fetchall()
        quotes = conn.execute(
            "SELECT * FROM quotes WHERE session_id=? ORDER BY version DESC LIMIT 1",
            (session_id,),
        ).fetchall()
        escalations = conn.execute(
            "SELECT * FROM escalation_queue WHERE session_id=?",
            (session_id,),
        ).fetchall()

    wb = openpyxl.Workbook()
    ctx = json.loads(session["context_json"]) if session and session["context_json"] else {}
    specs = ctx.get("product_specs", {})

    # Sheet 1: Client Info
    ws = wb.active or wb.create_sheet("Client Info")
    ws.title = "Client Info"
    _add_header_row(ws, ["Field", "Value"])
    client_data = [
        ("Session ID", session_id),
        ("Timestamp", session["created_at"] if session else ""),
        ("Client Name", session["client_name"] if session else ""),
        ("Company", session["client_company"] if session else ""),
        ("Email", session["client_email"] if session else ""),
        ("Phone", session["client_phone"] if session else ""),
        ("Address", session["client_address"] if session else ""),
    ]
    for r, (f, v) in enumerate(client_data, 2):
        ws.cell(row=r, column=1, value=f)
        ws.cell(row=r, column=2, value=v)

    # Sheet 2: Product Specs
    ws2 = wb.create_sheet("Product Specs")
    _add_header_row(ws2, ["Field", "Value"])
    spec_data = [
        ("Product Name", specs.get("product_name", "")),
        ("Product Type", specs.get("product_type", "")),
        ("Serving Size", specs.get("serving_size", "")),
        ("Servings/Unit", specs.get("servings_per_unit", "")),
        ("Total Count", specs.get("total_count", "")),
        ("Capsule Type", specs.get("capsule_type", "")),
        ("Order Quantity", specs.get("order_quantity", "")),
    ]
    for r, (f, v) in enumerate(spec_data, 2):
        ws2.cell(row=r, column=1, value=f)
        ws2.cell(row=r, column=2, value=v)

    # Sheet 3: Ingredients
    ws3 = wb.create_sheet("Ingredients")
    _add_header_row(ws3, ["Ingredient", "Source Tab", "Supplier", "mg/serving", "Label Claim", "UOM", "Unit Cost", "Total Cost"])
    for r, ing in enumerate(ingredients, 2):
        ws3.cell(row=r, column=1, value=ing["ingredient_name"])
        ws3.cell(row=r, column=2, value=ing["cost_source"])
        ws3.cell(row=r, column=4, value=ing["mg_per_serving"])
        ws3.cell(row=r, column=5, value=ing["label_claim"])
        ws3.cell(row=r, column=6, value=ing["uom"])
        ws3.cell(row=r, column=7, value=ing["unit_cost"])

    # Sheet 4: Pricing Summary
    ws4 = wb.create_sheet("Pricing Summary")
    _add_header_row(ws4, ["Line Item", "Low Estimate", "Mid Estimate", "High Estimate", "Data Source"])
    if quotes:
        q = quotes[0]
        pricing_rows = [
            ("Ingredients", q["ingredient_cost_low"], q["ingredient_cost_mid"], q["ingredient_cost_high"], "DB"),
            ("Machine Wear", q["machine_cost_low"], q["machine_cost_mid"], q["machine_cost_high"], "Config"),
            ("Labor", q["labor_cost_low"], q["labor_cost_mid"], q["labor_cost_high"], "Config"),
            ("Packaging", q["packaging_cost_low"], q["packaging_cost_mid"], q["packaging_cost_high"], "DB/Default"),
            ("Transportation", q["transport_cost_low"], q["transport_cost_mid"], q["transport_cost_high"], "Rate table"),
            ("TOTAL (with margin)", q["total_low"], q["total_mid"], q["total_high"], f"Margin: {q['margin_pct']}"),
        ]
        for r, (item, low, mid, high, src) in enumerate(pricing_rows, 2):
            ws4.cell(row=r, column=1, value=item)
            ws4.cell(row=r, column=2, value=low)
            ws4.cell(row=r, column=3, value=mid)
            ws4.cell(row=r, column=4, value=high)
            ws4.cell(row=r, column=5, value=src)

    # Sheet 5: Chat Transcript
    ws5 = wb.create_sheet("Chat Transcript")
    _add_header_row(ws5, ["Timestamp", "Speaker", "Phase", "Message"])
    for r, msg in enumerate(messages, 2):
        ws5.cell(row=r, column=1, value=msg["timestamp"])
        ws5.cell(row=r, column=2, value=msg["role"])
        ws5.cell(row=r, column=3, value=msg["phase"])
        ws5.cell(row=r, column=4, value=(msg["content"] or "")[:500])

    # Sheet 6: Escalated Items
    ws6 = wb.create_sheet("Escalated Items")
    _add_header_row(ws6, ["Item", "Source", "Requested Date", "Status", "Admin Response", "Confirmed Price"])
    for r, esc in enumerate(escalations, 2):
        ws6.cell(row=r, column=1, value=esc["item_requested"])
        ws6.cell(row=r, column=2, value=esc["source"])
        ws6.cell(row=r, column=3, value=esc["created_at"])
        ws6.cell(row=r, column=4, value=esc["status"])
        ws6.cell(row=r, column=5, value=esc["admin_notes"])
        ws6.cell(row=r, column=6, value=esc["confirmed_price"])

    filename = f"client_record_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = CLIENT_RECORDS_DIR / filename
    wb.save(str(filepath))
    logger.info("Exported client record to %s", filepath)
    return str(filepath)
