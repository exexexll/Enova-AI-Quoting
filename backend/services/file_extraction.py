"""Extract content from uploaded files (PDF, Excel, images).

- PDF: text extraction via pdfplumber
- Excel: cell data extraction via openpyxl
- Images: GPT vision API for label/formulation analysis
"""
from __future__ import annotations

import base64
import logging
from pathlib import Path
from typing import Optional

from backend.config import OPENAI_API_KEY, OPENAI_MODEL

logger = logging.getLogger(__name__)

MAX_TEXT_CHARS = 12000


def extract_file_content(filepath: str, content_type: str | None = None) -> str:
    """Extract readable content from a file. Returns extracted text or empty string."""
    path = Path(filepath)
    if not path.exists():
        return ""

    suffix = path.suffix.lower()
    ct = (content_type or "").lower()

    try:
        if suffix == ".pdf" or "pdf" in ct:
            return _extract_pdf(path)
        elif suffix in (".xlsx", ".xls") or "spreadsheet" in ct or "excel" in ct:
            return _extract_excel(path)
        elif suffix in (".png", ".jpg", ".jpeg", ".gif", ".webp") or ct.startswith("image/"):
            return _extract_image_vision(path, ct or f"image/{suffix.lstrip('.')}")
        elif suffix in (".txt", ".csv", ".json", ".md"):
            text = path.read_text(errors="replace")[:MAX_TEXT_CHARS]
            return text
    except Exception as e:
        logger.warning("File extraction failed for %s: %s", filepath, e)

    return ""


def _extract_pdf(path: Path) -> str:
    """Extract text from all pages of a PDF."""
    try:
        import pdfplumber
    except ImportError:
        logger.warning("pdfplumber not installed; PDF extraction unavailable")
        return ""

    text_parts: list[str] = []
    with pdfplumber.open(str(path)) as pdf:
        for i, page in enumerate(pdf.pages):
            page_text = page.extract_text() or ""
            tables = page.extract_tables()

            if page_text.strip():
                text_parts.append(f"--- Page {i + 1} ---\n{page_text.strip()}")

            for t_idx, table in enumerate(tables):
                if not table:
                    continue
                rows = []
                for row in table:
                    cells = [str(c).strip() if c else "" for c in row]
                    rows.append(" | ".join(cells))
                if rows:
                    header = rows[0]
                    sep = " | ".join(["---"] * len(table[0])) if table[0] else "---"
                    body = "\n".join(rows[1:])
                    text_parts.append(
                        f"[Table {t_idx + 1} on page {i + 1}]\n{header}\n{sep}\n{body}"
                    )

    combined = "\n\n".join(text_parts)
    if len(combined) > MAX_TEXT_CHARS:
        combined = combined[:MAX_TEXT_CHARS] + "\n\n[... content truncated ...]"

    return combined


def _extract_excel(path: Path) -> str:
    """Extract data from all sheets of an Excel file."""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    text_parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))

        if rows:
            header = rows[0]
            sep = " | ".join(["---"] * len(rows[0].split(" | ")))
            body = "\n".join(rows[1:30])
            text_parts.append(
                f"[Sheet: {sheet_name}] ({len(rows)} rows)\n{header}\n{sep}\n{body}"
            )
            if len(rows) > 30:
                text_parts.append(f"[... {len(rows) - 30} more rows ...]")

    wb.close()

    combined = "\n\n".join(text_parts)
    if len(combined) > MAX_TEXT_CHARS:
        combined = combined[:MAX_TEXT_CHARS] + "\n\n[... content truncated ...]"

    return combined


def _extract_image_vision(path: Path, content_type: str) -> str:
    """Use GPT vision to analyze an image (label, formulation, ingredient list)."""
    if not OPENAI_API_KEY:
        return "[Image uploaded but no OpenAI API key for vision analysis]"

    try:
        from openai import OpenAI

        image_data = base64.b64encode(path.read_bytes()).decode("utf-8")
        mime = content_type if "/" in content_type else f"image/{path.suffix.lstrip('.')}"

        client = OpenAI(api_key=OPENAI_API_KEY)
        response = client.chat.completions.create(
            model=OPENAI_MODEL,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are a supplement industry expert analyzing an uploaded image. "
                        "Extract ALL information you can see: ingredient names, dosages, "
                        "serving sizes, product name, supplement facts panel data, "
                        "formulation details, pricing, quantities, and any other relevant "
                        "product specifications. Present the data in a structured format "
                        "with markdown tables where appropriate. Be thorough and precise."
                    ),
                },
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime};base64,{image_data}",
                                "detail": "high",
                            },
                        },
                        {
                            "type": "text",
                            "text": "Analyze this image and extract all supplement/product information visible.",
                        },
                    ],
                },
            ],
            max_completion_tokens=2000,
        )

        result = response.choices[0].message.content or ""
        logger.info("Vision analysis completed for %s (%d chars)", path.name, len(result))
        return result

    except Exception as e:
        logger.warning("Vision analysis failed for %s: %s", path.name, e)
        return f"[Image uploaded but vision analysis failed: {e}]"
